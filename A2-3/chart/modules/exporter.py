# ============================================================
# A 담당 · 내보내기  (원래 B 영역이었으나 A로 이관 · 명세 4.9)
#
# 이 파일은 DB를 모른다. 조회는 A의 get_reviews_for_export() 가 하고
# 여기서는 받은 리스트를 파일로 쓰기만 한다.
# 조회 조건이 늘어도 이 파일은 안 바뀐다.
#
# ------------------------------------------------------------
# 가장 중요한 것: 선택 필드가 None 인 리뷰를 처리해야 한다
#
#   rating / review_date / skin_type / sentiment / confidence 는
#   전부 None 일 수 있다. (명세 4.2 - 선택 필드)
#
#   int(record["rating"]) 을 무조건 부르면
#       TypeError: int() argument must be a string ... not 'NoneType'
#   로 터진다. 별점을 안 남긴 리뷰 한 건 때문에 export 전체가 죽는다.
#
#   csv.DictWriter 는 None 을 빈 문자열로 써주지만,
#   숫자 포맷팅(f"{confidence:.2f}")은 None 에서 터진다.
#   그래서 포맷팅 전에 전부 한 번 걸러낸다.
# ============================================================

import csv
import json
from pathlib import Path

from modules.logger import get_logger


logger = get_logger("exporter")


SUPPORTED_FORMATS = ("csv", "jsonl", "xlsx")

# 내보낼 컬럼과 헤더. 순서를 여기서 한 번만 정한다.
COLUMNS = [
    ("id", "id"),
    ("product_name", "제품명"),
    ("review_text", "리뷰"),
    ("rating", "별점"),
    ("review_date", "작성일"),
    ("skin_type", "피부타입"),
    ("sentiment", "감정"),
    ("confidence", "확신도"),
    ("language", "언어"),
    ("model", "분석모델"),
    ("analyzed_at", "분석일시"),
]

FIELDS = [key for key, _ in COLUMNS]
HEADERS = [header for _, header in COLUMNS]


def normalize(record):
    """
    레코드 1건을 내보내기용으로 정리한다.

    None 은 빈 문자열로, confidence 는 소수 2자리로.
    여기서 한 번만 처리하면 세 포맷이 같은 값을 쓴다.
    csv 와 xlsx 의 숫자가 서로 다르면 나중에 대조할 때 사람이 헷갈린다.
    """

    row = {}

    for key in FIELDS:
        value = record.get(key)

        if value is None:
            row[key] = ""

        elif key == "confidence":
            # float 이 아닐 수도 있다. 문자열이 들어와도 죽지 않게 한다.
            try:
                row[key] = round(float(value), 2)
            except (TypeError, ValueError):
                row[key] = ""

        elif key == "rating":
            try:
                row[key] = int(value)
            except (TypeError, ValueError):
                row[key] = ""

        else:
            row[key] = value

    return row


def export_csv(records, output_path):
    """
    CSV. Excel 이 한글을 깨뜨리지 않도록 utf-8-sig 로 쓴다.

    utf-8 로 쓰면 Excel 이 BOM 이 없는 파일을 cp949 로 읽어서
    한글이 전부 깨진다. 팀에서 파일을 열어보는 사람은 대개 Excel 을 쓴다.
    newline="" 은 윈도우에서 빈 줄이 하나씩 끼는 것을 막는다.
    """

    with open(output_path, "w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(HEADERS)

        for record in records:
            row = normalize(record)
            writer.writerow([row[key] for key in FIELDS])

    return output_path


def export_jsonl(records, output_path):
    """
    JSONL. 한 줄에 한 건이라 다른 스크립트가 스트리밍으로 읽기 좋다.

    여기서는 None 을 빈 문자열로 바꾸지 않고 그대로 null 로 둔다.
    JSON 을 읽는 쪽은 프로그램이고, "값이 없음"과 "빈 문자열"은 다르다.
    (CSV 는 사람이 보는 포맷이라 빈 칸이 낫다)
    """

    with open(output_path, "w", encoding="utf-8") as file:

        for record in records:
            row = {key: record.get(key) for key in FIELDS}

            if row.get("confidence") is not None:

                try:
                    row["confidence"] = round(float(row["confidence"]), 2)
                except (TypeError, ValueError):
                    row["confidence"] = None

            file.write(json.dumps(row, ensure_ascii=False) + "\n")

    return output_path


def export_xlsx(records, output_path):
    """
    XLSX. openpyxl 이 없으면 CSV 로 떨어진다.

    포맷 하나 때문에 명령 전체가 실패하는 것보다,
    같은 데이터를 열 수 있는 형태로 주고 경고를 남기는 편이 낫다.
    """

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

    except ImportError:
        logger.warning(
            "openpyxl 이 없어 CSV 로 내보냅니다. "
            "설치: pip install openpyxl"
        )

        fallback = Path(output_path).with_suffix(".csv")

        return export_csv(records, fallback)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "reviews"

    sheet.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="F0EFEC")

    for index in range(1, len(HEADERS) + 1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for record in records:
        row = normalize(record)
        sheet.append([row[key] for key in FIELDS])

    # 열 너비. 리뷰 본문만 넓게 잡고 나머지는 헤더에 맞춘다.
    widths = {"id": 6, "product_name": 18, "review_text": 60, "rating": 7,
              "review_date": 12, "skin_type": 10, "sentiment": 10,
              "confidence": 9, "language": 7, "model": 18, "analyzed_at": 20}

    for index, key in enumerate(FIELDS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths[key]

    review_column = FIELDS.index("review_text") + 1

    for row_cells in sheet.iter_rows(min_row=2, min_col=review_column,
                                     max_col=review_column):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 첫 행 고정 + 필터. 99행을 스크롤하면 헤더가 사라진다.
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    workbook.save(output_path)

    return output_path


WRITERS = {
    "csv": export_csv,
    "jsonl": export_jsonl,
    "xlsx": export_xlsx,
}


def export_reviews(records, fmt, output_path):
    """
    리뷰 리스트를 파일로 내보내고 실제 저장 경로를 돌려준다.

    반환값이 '실제 경로' 인 이유
      xlsx 요청이 openpyxl 부재로 csv 로 떨어질 수 있다.
      호출부가 요청한 경로를 그대로 출력하면 실제로 없는 파일을 안내하게 된다.
    """

    fmt = str(fmt).lower().lstrip(".")

    if fmt not in WRITERS:
        raise ValueError(
            f"지원하지 않는 형식입니다: {fmt} "
            f"(가능: {', '.join(SUPPORTED_FORMATS)})"
        )

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    saved = WRITERS[fmt](records, output_path)

    logger.info("%d건 내보내기 완료: %s", len(records), saved)

    return str(saved)
