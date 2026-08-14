# ============================================================
# A 담당 · 외부 폴더(source/, prompt/) 호출 창구
#
# 이전 adapters.py 를 대체한다. 539줄 -> 이 파일.
#
# 무엇이 사라졌나
#   C(민규)와 합의해서 C 코드를 조금 고쳤더니, A쪽에서 하던 변환이
#   대부분 필요 없어졌다. 없앤 것들:
#
#     · id 매핑 73줄      -> C가 id 를 그대로 들고 다녀서 아예 불필요
#     · 부분 실패 흡수     -> C가 failed_ids 를 함께 돌려준다
#     · 배치 나누기        -> C가 직접 배치로 부른다 (아래 주석)
#     · stats 평면화 77줄  -> 리포트 형식은 리포트를 만드는 C가 정한다
#     · chart_paths 변환   -> C가 dict 도 받는다. 상대경로 계산만 A가 3줄
#     · insights None 방어 -> C가 기본값을 넣는다
#     · 차트/export 통과   -> A가 A를 부르는 껍데기였다. 직접 import 한다
#     · mock 분기         -> 아래 [2026-08-12] 참고
#
# 무엇이 남았나
#   · B(세인) 래퍼 — B의 함수는 파일 경로를 받고 파일을 쓴다.
#     A는 DB에 넣어야 해서 그 사이를 이어야 한다.
#     B와는 아직 이 합의를 하지 않았고, 지금 형태로도 잘 돌아간다.
#   · language 채우기 — C가 안 주는데 DB 컬럼에는 있다.
#
# [2026-08-12] mock 분기 제거
#   C의 Gemini 연결이 끝나서 대역(modules/mock_ai.py)을 걷어냈다.
#   대역을 남겨두면 "지금 도는 게 진짜 분석인가 대역인가" 를 실행할 때마다
#   플래그로 확인해야 하고, 규칙 기반 결과가 DB에 섞여 들어가도
#   model 컬럼을 열어보기 전까지는 아무도 모른다.
#   자가 점검과 테스트가 쓰던 대역은 각각 그 파일 안으로 옮겼다.
#   덕분에 analyzer() / extractor() 가 config 를 받을 이유도 없어졌다.
#
# 원칙은 그대로다. 의존 방향은 A -> B, A -> C 한 방향이다.
# B/C 는 A를 import 하지 않는다. 그래야 두 사람이 자기 폴더만으로
# 테스트할 수 있다.
# ============================================================

import csv
import os
import re
import tempfile
from pathlib import Path

from modules.logger import get_logger
from modules.paths import (DATA_CLEAN_DIR, DATA_RAW_DIR, ModuleNotProvided,
                           load_b_cleaner, load_b_importer, load_c_analyzer,
                           load_c_extractor, load_c_reporter)


logger = get_logger("bridge")

HANGUL = re.compile(r"[가-힣]")


def detect_language(text):
    """
    아주 단순한 언어 판별.

    C는 language 를 돌려주지 않는데 DB 스키마와 show 출력에는 칸이 있다.
    이것 하나 때문에 모델을 한 번 더 부르는 건 낭비라 한글 포함 여부만 본다.
    계약상 nullable 이라 틀려도 치명적이지 않다.
    """

    if not text:
        return None

    return "ko" if HANGUL.search(text) else "en"


# ============================================================
# B (세인) — source/src
# ============================================================
#
# B의 두 함수는 '파일 경로를 받아 파일을 쓰고 리스트를 돌려주는' 형태다.
# A는 DB에 넣어야 하니 경로를 만들어 주고 결과 리스트만 받는다.
# B가 만든 raw/clean JSONL 은 그대로 남아서, B가 혼자 돌릴 때와
# 산출물 위치가 달라지지 않는다.


EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx"}


def xlsx_to_csv(file_path, output_path):
    """
    Excel 첫 시트를 CSV 로 옮긴다.

    왜 B의 importer 를 고치지 않고 여기서 변환하나
      B의 importer.py 는 세인님 파일이라 A가 손대지 않는다.
      그런데 '어떤 파일 형식을 받는가' 는 CLI 의 입출력 범위 문제라
      A가 책임지는 게 맞다. 그래서 Excel 을 CSV 로 한 번 눕혀
      B에게는 지금까지와 똑같이 CSV 경로만 넘긴다.

      B 코드는 한 줄도 바뀌지 않고, 정제 규칙도 한 곳에 그대로 남는다.
      나중에 세인님이 xlsx 를 직접 읽게 되면 이 함수만 지우면 된다.

    첫 시트만 읽는 이유
      시트가 여러 장인 파일에서 어느 시트가 리뷰인지는 A가 알 수 없다.
      두 번째 시트가 필요하면 그건 합의가 필요한 이야기라
      지금은 첫 시트로 못 박고 로그에 시트 이름을 남긴다.

    빈 행은 건너뛴다. Excel 은 한 번 값이 들어갔던 셀을 지워도
    행 자체는 남겨서, 그대로 옮기면 빈 줄이 수백 개 딸려온다.
    """

    try:
        from openpyxl import load_workbook

    except ImportError as error:
        raise ValueError(
            f"Excel 파일을 읽으려면 openpyxl 이 필요합니다: {error}\n"
            f"설치: pip install -r chart/requirements.txt"
        ) from error

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    written = 0

    with open(output_path, "w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)

        for row in sheet.iter_rows(values_only=True):

            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            writer.writerow(
                ["" if cell is None else str(cell).strip() for cell in row]
            )
            written += 1

    workbook.close()

    if written <= 1:
        raise ValueError(
            f"Excel 에서 읽은 데이터 행이 없습니다 "
            f"(시트: {sheet.title}, 읽은 줄: {written}). "
            f"첫 시트에 헤더와 데이터가 있는지 확인하세요."
        )

    logger.info(
        "Excel -> CSV 변환: %s 시트 %d줄 (헤더 포함)", sheet.title, written
    )

    return output_path


def import_raw_file(file_path, output_path=None):
    """
    [B] CSV/Excel -> raw JSONL. (원본 dict 리스트, 저장 경로) 반환.

    Excel 이면 임시 CSV 로 눕힌 뒤 B에게 넘긴다. 위 xlsx_to_csv 주석 참고.
    """

    output_path = output_path or (DATA_RAW_DIR / "reviews.jsonl")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    source = Path(file_path)

    if not source.exists():
        # B도 FileNotFoundError 를 던지지만, Excel 변환이 먼저 터지면
        # openpyxl 의 메시지가 나가서 원인이 흐려진다.
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {source}")

    temporary = None

    try:
        if source.suffix.lower() in EXCEL_SUFFIXES:
            handle, temporary = tempfile.mkstemp(suffix=".csv", text=True)
            os.close(handle)
            xlsx_to_csv(source, temporary)
            source = Path(temporary)

        # B는 FileNotFoundError / ValueError 를 던진다.
        # 그대로 올려보내고 main.py 가 종료 코드로 바꾼다.
        rows = load_b_importer().import_reviews(
            str(source), str(output_path)
        )

    finally:
        if temporary:
            Path(temporary).unlink(missing_ok=True)

    logger.info("B importer: %d건 읽음 -> %s", len(rows), output_path)

    return rows, output_path


def clean_raw_file(input_path=None, output_path=None):
    """[B] raw JSONL -> clean JSONL. (정제 dict 리스트, 저장 경로) 반환."""

    input_path = input_path or (DATA_RAW_DIR / "reviews.jsonl")
    output_path = output_path or (DATA_CLEAN_DIR / "reviews.jsonl")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    records = load_b_cleaner().clean_reviews(str(input_path), str(output_path))

    logger.info("B cleaner: %d건 정제 -> %s", len(records), output_path)

    return records, output_path


def write_raw_jsonl(records, output_path):
    """
    DB의 raw 저장소를 B가 읽을 수 있는 JSONL 로 되돌린다.

    clean 명령(원본 CSV 없이 재정제)에서 쓴다.
    B의 clean_reviews 는 파일 경로만 받으므로 파일을 한 번 거쳐야 한다.
    """

    import json

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as file:
        for record in records:
            file.write(json.dumps(record, ensure_ascii=False) + "\n")

    return output_path


# ============================================================
# C (민규) — prompt
# ============================================================
#
# A가 배치를 나누지 않는 이유
#   C의 analyze_reviews() 가 한 프롬프트에 여러 건을 넣어 부르고,
#   실패하면 스스로 반으로 쪼개 다시 시도한다.
#   A가 앞에서 또 자르면 그 재시도 폭만 좁아질 뿐 이득이 없다.
#   진짜 배치는 C 영역에서만 가능하다.


def analyzer():
    """감정 분석 모듈 (prompt/analyzer.py)."""

    return load_c_analyzer()


def extractor():
    """인사이트 추출 모듈 (prompt/extractor.py)."""

    return load_c_extractor()


def reporter():
    """리포트 모듈 (prompt/reporter.py). 모델을 부르지 않는다."""

    return load_c_reporter()


__all__ = [
    "ModuleNotProvided",
    "detect_language",
    "xlsx_to_csv",
    "import_raw_file",
    "clean_raw_file",
    "write_raw_jsonl",
    "analyzer",
    "extractor",
    "reporter",
]
